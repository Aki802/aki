import pyautogui as ag
import time
import pytesseract
import pyperclip as pc
import numpy
import cv2
import flet as ft
import openpyxl as xl
from pathlib import Path
import win32com.client
import os
import xlwings as xw
import pythoncom
import pickle
import keyboard

Keys = []
SS_count = [1]

#指定した起動中のアプリ選択
def move(key):
    KEY=str(key)
    ag.hotkey("win", KEY)

#アプリのカーソル位置取得
def operation():
    P1 = ag.position()
    time.sleep(2)
    while True:
        P2 = ag.position()
        if P1==P2:
            break
        time.sleep(1)
        P1 = P2
    return P1

#文字入力
def moji(mozi):
    pc.copy(mozi)
    ag.hotkey('ctrl', 'v')

#マウス右クリック
def Rclick():
    ag.rightClick()

#マウス左クリック
def Lclick():
    ag.leftClick()

#カーソル移動
def move_carsol(posi):
    ag.moveTo(posi)

#ドラッグドロップ
def drag_drop(Pos1, Pos2):
    ag.moveTo(Pos1)
    ag.mouseDown()
    ag.moveTo(Pos2, duration=0.5)
    ag.mouseUp()

#エンター
def Enter():
    ag.press('enter')

#スクロール
def scroll(ryou):
    ag.hscroll(ryou)

#コピー
def copy():
    ag.hotkey("ctrl", "C")

#貼り付け
def paste():
    ag.hotkey("ctrl", "V")
    
#全画面表示
def fullscreen():
    ag.hotkey("win", "up")

#スクリーンショット
def screenShot():
    folder_path = "evidence"
    SS = len(SS_count)
    SSS = str(SS)
    ii = 1
    full_file_path = os.path.join(folder_path, "Screenshot"+SSS+".png")
    while os.path.exists(full_file_path):
        iii = str(ii)
        full_file_path = os.path.join(folder_path, "Screenshot"+SSS+"("+iii+")"+".png")
        ii = ii+1
    SS_count.append(SS+1)
    scsh = ag.screenshot()
    scsh.save(full_file_path)



#キーコンフィグ
def keycon(keys):
    ag.hotkey(keys)



#Excelを開く
def xls_open(xlpath):
    wb = xl.load_workbook(xlpath)
    return wb

#新しいExcelを作成
def new_xls(Xname):
    wb = xl.Workbook()
    wb.save(Xname)
    return wb

#シート選択
def xls_sheet(sheet, wb):
    ws = wb[sheet]
    return ws

#Excelに書き込む
def xls_write(X, y, mo, wb, ws, xlpath):
    Y = ord(y)
    Y = Y-64
    ws.cell(row=X, column=Y, value=mo)
    wb.save(xlpath)

#マクロ読み込み
def read_vba_macro(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        vba_code = file.read()
    return vba_code
        

#VBAマクロ追加
def add_vba_macro_excel(excel_file_path, vba_macro, module_name='Module1'):
    # Excelを起動
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    
    # 指定したExcelファイルを開く
    workbook = excel.Workbooks.Open(excel_file_path)
    
    # モジュールが既に存在するかチェック
    vb_component = None
    for vb_comp in workbook.VBProject.VBComponents:
        if vb_comp.Name == module_name:
            vb_component = vb_comp
            break
    
    if vb_component:
        # 既存のモジュールがある場合、そのコードを更新
        vb_component.CodeModule.DeleteLines(1, vb_component.CodeModule.CountOfLines)
        vb_component.CodeModule.AddFromString(vba_macro)
    else:
        # 新しいモジュールを追加
        vb_component = workbook.VBProject.VBComponents.Add(1)  # 1:標準モジュール
        vb_component.Name = module_name
        vb_component.CodeModule.AddFromString(vba_macro)
    
    workbook.Save()
    workbook.Close()

#VBA起動
def ex_vba_macro(excel_file_path):

    app = xw.App()

    wb = app.books.open(excel_file_path)

    pg = wb.macro("A1_in_test")

    pg()

    wb.save(excel_file_path)

    wb.close()
    app.quit()


#setupファイル出力
def export_setup(count, setname):
    with open(setname, 'wb') as f:
        pickle.dump(count, f)




#UI
def main(page: ft.Page):
    page.title = "AutoGUI"
    page.bgcolor = ft.colors.WHITE
    page.window_title_bar_hidden = True
    page.window_title_bar_buttons_hidden = True
    


    locate = []
    count = []
    ryo = []
    

    if os.path.exists('setup/setup.pkl') and os.path.getsize('setup/setup.pkl') > 0:
        with open('setup/setup.pkl', 'rb') as f:
            count = pickle.load(f)
    else:
        print("setupfileなし")

    
    

    #位置情報の追加
    def carsol_posi_clicked(e):
        comment.value = "取得したい位置にマウスを移動し1秒待機してください"
        page.update()
        posi = operation()
        locate.append(posi)
        kazu = len(locate) - 1
        locate_row.controls.append(ft.Text(kazu))
        locate_row.controls.append(ft.Text("="))
        locate_row.controls.append(ft.Text(posi))
        locate_row.controls.append(ft.Text("   "))
        new_value = len(ryo)
        ryo.append(new_value)
        update_dd()
        comment.value = "位置情報を追加しました"
        page.update()
        

    def W_button_click(e):
        T = second_W.value

        #チャートに追加
        count.append('TIME')
        count.append(T)
        count_column.controls.append(ft.Row([ft.Text(T), ft.Text("秒待機")]))
        comment.value = "チャートを追加しました"
        page.update()


    #動作系↓

    def A_submi_click(e):
        tool = A_number.value

        #チャートに追加
        count.append('TOOL')
        count.append(tool)
        count_column.controls.append(ft.Row([ft.Text("ツールバーの"), ft.Text(tool), ft.Text("番目のアプリを選択")]))
        comment.value = "チャートを追加しました"
        page.update()

    def mozi_button_click(e):
        Mozi = mozi_field.value

        #チャートに追加
        count.append('MOZI')
        count.append(Mozi)
        count_column.controls.append(ft.Row([ft.Text("文字列「"), ft.Text(Mozi), ft.Text("」を入力")]))
        comment.value = "チャートを追加しました"
        page.update()

    def posi_button_click(e):
        Posi = move_posi.value
        Posi = int(Posi)
        Posi = locate[Posi]

        #チャートに追加
        count.append('POSI')
        count.append(Posi)
        count_column.controls.append(ft.Row([ft.Text("カーソルを"), ft.Text(Posi), ft.Text("に移動")]))
        comment.value = "チャートを追加しました"
        page.update()

    def lclick_button_click(e):
        count.append('LCLICK')
        count_column.controls.append(ft.Row([ft.Text("左クリック")]))
        comment.value = "チャートを追加しました"
        page.update()

    def rclick_button_click(e):
        count.append('RCLICK')
        count_column.controls.append(ft.Row([ft.Text("右クリック")]))
        comment.value = "チャートを追加しました"
        page.update()

    def drdr_button_click(e):
        stposi = start_posi.value
        stposi = int(stposi)
        stposi = locate[stposi]

        enposi = end_posi.value
        enposi = int(enposi)
        enposi = locate[enposi]

        #チャートに追加
        count.append('DRDR')
        count.append(stposi)
        count.append(enposi)
        count_column.controls.append(ft.Row([ft.Text(stposi), ft.Text("から"), ft.Text(enposi), ft.Text("にドラッグドロップ")]))
        comment.value = "チャートを追加しました"
        page.update()

    def copy_click(e):
        count.append('COPY')
        count_column.controls.append(ft.Row([ft.Text("コピー")]))
        comment.value = "チャートを追加しました"
        page.update()

    def paste_click(e):
        count.append('PASTE')
        count_column.controls.append(ft.Row([ft.Text("ペースト")]))
        comment.value = "チャートを追加しました"
        page.update()

    def enter_click(e):
        count.append('ENTER')
        count_column.controls.append(ft.Row([ft.Text("エンター")]))
        comment.value = "チャートを追加しました"
        page.update()

    def full_button_click(e):
        count.append('FULL')
        count_column.controls.append(ft.Row([ft.Text("全画面表示")]))
        comment.value = "チャートを追加しました"
        page.update()
    
    def screenshot_click(e):
        #チャートに追加
        count.append('SHOT')
        count_column.controls.append(ft.Row([ft.Text("スクリーンショット")]))
        comment.value = "チャートを追加しました"
        page.update()
        

        


    def key_in(e, KEYS):
        comment.value = "キーを入力してください"
        page.update()
        keyin = keyboard.read_key()
        if keyin=="left windows":
            keyin = "win"
        elif keyin=="right ctrl":
            keyin = "ctrl"
        elif keyin=="right shift":
            keyin = "shift"
        Keys.append(keyin)
        keycons = ", ".join(KEYS)
        key_field.value = keycons
        comment.value = "キーを追加しました"
        page.update()

    def key_button_click(e, KEYS):
        keycons = ", ".join(KEYS)
        KEYSBU = Keys.copy()

        #チャートに追加
        count.append('KEYCON')
        count.append(KEYSBU)
        count_column.controls.append(ft.Row([ft.Text(value=keycons), ft.Text("のキーを入力")]))
        comment.value = "チャートを追加しました"


        #Keysリセット
        Keys.clear()
        key_field.value = ""
        page.update()






    #Excel関係↓
    #新しいエクセル作成
    def newxl_ST(e):
        name = newxl_name.value

        #チャートに追加
        count.append('NEWXL')
        count.append(name)
        count_column.controls.append(ft.Row([ft.Text(name), ft.Text(".xlsxを作成")]))
        comment.value = "チャートを追加しました"
        page.update()

    #VBA取得
    def VBA_select_click(e: ft.FilePickerResultEvent):
        if e.files:
            count.append('VBA')
            count.append(e.files[0].path)
            count_column.controls.append(ft.Row([ft.Text(e.files[0].path), ft.Text("を")]))
            VBA_select.visible = False
            excel_select.visible = True
            comment.value = "Excelを追加してください"
            page.update()




    def excel_select_click(e: ft.FilePickerResultEvent):
        if e.files:
            count.append(e.files[0].path)
            count_column.controls.append(ft.Row([ft.Text(e.files[0].path), ft.Text("に適応")]))
            VBA_select.visible = True
            excel_select.visible = False
            comment.value = "チャートを追加しました"
            page.update()

    #importの追加ーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーー

    def import_add(i):
        while True:
            if i >= len(count):
                page.update()
                break
            elif count[i]=="TIME":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text(count[i]), ft.Text("秒待機")]))
                i=i+1
            elif count[i]=="TOOL":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text("ツールバーの"), ft.Text(count[i]), ft.Text("番目のアプリを選択")]))
                i=i+1
            elif count[i]=="MOZI":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text("文字列「"), ft.Text(count[i]), ft.Text("」を入力")]))
                i=i+1
            elif count[i]=="POSI":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text("カーソルを"), ft.Text(count[i]), ft.Text("に移動")]))
                i=i+1
            elif count[i]=="LCLICK":
                count_column.controls.append(ft.Row([ft.Text("左クリック")]))
                i=i+1
            elif count[i]=="RCLICK":
                count_column.controls.append(ft.Row([ft.Text("右クリック")]))
                i=i+1
            elif count[i]=="ENTER":
                count_column.controls.append(ft.Row([ft.Text("エンター")]))
                i=i+1
            elif count[i]=="DRDR":
                i=i+1
                stdr=count[i]
                i=i+1
                endr=count[i]
                count_column.controls.append(ft.Row([ft.Text(stdr), ft.Text("から"), ft.Text(endr), ft.Text("にドラッグドロップ")]))
                i=i+1
            elif count[i]=="COPY":
                count_column.controls.append(ft.Row([ft.Text("コピー")]))
                i=i+1
            elif count[i]=="PASTE":
                count_column.controls.append(ft.Row([ft.Text("ペースト")]))
                i=i+1
            elif count[i]=="FULL":
                count_column.controls.append(ft.Row([ft.Text("全画面表示")]))
                i=i+1
            elif count[i]=="SHOT":
                count_column.controls.append(ft.Row([ft.Text("スクリーンショット")]))
                i=i+1
            elif count[i]=="KEYCON":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text(count[i]), ft.Text("のキーを入力")]))
                i=i+1
            elif count[i]=="NEWXL":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text(count[i]), ft.Text(".xlsxを作成")]))
                i=i+1
            elif count[i]=="VBA":
                i=i+1
                macro_path = count[i]
                i=i+1
                excel_path = count[i]
                count_column.controls.append(ft.Row([ft.Text(macro_path), ft.Text("を")]))
                count_column.controls.append(ft.Row([ft.Text(excel_path), ft.Text("に適応")]))
                i=i+1
            else:
                page.update()
                break

        
    def import_select_click(e: ft.FilePickerResultEvent):
        FilePath = e.files[0].path
        BUNEW = []
        i = len(count)
        with open(FilePath, 'rb') as f:
            BUNEW = pickle.load(f)
            count.extend(BUNEW)
        import_add(i)
        
    def txt_select_click(e: ft.FilePickerResultEvent):
        file_name = e.files[0].path
        txt_chart = []

        with open(file_name, 'r', encoding='utf-8') as file:
            for line in file:
                txt_chart.append(line.strip())

        folder_path = 'database'
        
        #それぞれのチャートを追加
        for j in txt_chart:
            i = len(count)
            BUPKL = []
            pkl_name = j+".pkl"
            full_file_path = os.path.join(folder_path, pkl_name)
            if os.path.exists(full_file_path) and os.path.getsize(full_file_path) > 0:
                with open(full_file_path, 'rb') as f:
                    BUPKL = pickle.load(f)
                    count.extend(BUPKL)
                    import_add(i)
            else:
                comment.value = f"「{j}」の手順がありません　追加してください"
                set_text.value = j
                page.update()
                break
            comment.value = "テキストからチャートを読み込みました"
            page.update()

        


    def clear_click(e):
        if count_column.controls:
            dast = count.pop()
            #追加したときに追記ーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーー
            while dast!="TIME" and dast!="TOOL" and dast!="MOZI" and dast!="POSI" and dast!="LCLICK" and dast!="RCLICK" and dast!="ENTER" and dast!="DRDR" and dast!="COPY" and dast!="PASTE" and dast!="FULL" and dast!="SHOT" and dast!="KEYCON" and dast!="NEWXL" and dast!="VBA":
                dast = count.pop()
            count_column.controls.pop()
            comment.value = "チャートを削除しました"
            page.update()

    def set_click(e):
        folder_path = 'database'
        set_name = set_text.value
        set_name = set_name+".pkl"
        full_file_path = os.path.join(folder_path, set_name)
        export_setup(count, full_file_path)
        set_text.value = ""
        comment.value = "databaseにセットアップファイルを作成しました"
        page.update()

            

    #起動フラグ　新しいエクセル作成(NEWXL,エクセルの名前)　VBA取得(VBA,マクロテキストのパス,エクセルのパス)ーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーー
    def start_button_click(e):
        i=0
        print(count)
        while True:
            if i >= len(count):
                comment.value = "チャートを実行しました"
                page.update()
                break
            elif count[i]=="TIME":
                i=i+1
                sr = int(count[i])
                time.sleep(sr)
                i=i+1
            elif count[i]=="TOOL":
                i=i+1
                move(count[i])
                i=i+1
            elif count[i]=="MOZI":
                i=i+1
                moji(count[i])
                i=i+1
            elif count[i]=="POSI":
                i=i+1
                move_carsol(count[i])
                i=i+1
            elif count[i]=="LCLICK":
                Lclick()
                i=i+1
            elif count[i]=="RCLICK":
                Rclick()
                i=i+1
            elif count[i]=="ENTER":
                Enter()
                i=i+1
            elif count[i]=="DRDR":
                i=i+1
                stdr=count[i]
                i=i+1
                endr=count[i]
                drag_drop(stdr, endr)
                i=i+1
            elif count[i]=="COPY":
                copy()
                i=i+1
            elif count[i]=="PASTE":
                paste()
                i=i+1
            elif count[i]=="FULL":
                fullscreen()
                i=i+1
            elif count[i]=="SHOT":
                screenShot()
                i=i+1
            elif count[i]=="KEYCON":
                i=i+1
                keycon(count[i])
                i=i+1
            elif count[i]=="NEWXL":
                i=i+1
                Xname = count[i]+".xlsx"
                cd = Path.cwd()
                path = cd / Xname
                new_xls(path)
                print(count[i])
                i=i+1
            elif count[i]=="VBA":
                i=i+1
                macro_path = count[i]
                i=i+1
                excel_path = count[i]

                print(macro_path)
                print(excel_path)
                
                # VBAマクロの読み込み
                vba_macro = read_vba_macro(macro_path)

                pythoncom.CoInitialize()  # Excelを起動する前にこれを呼び出す
                # VBAマクロをExcelに追加
                add_vba_macro_excel(excel_path, vba_macro)

                #  VBAマクロを起動
                ex_vba_macro(excel_path)
                pythoncom.CoUninitialize()  # Excelを終了した後はこれを呼び出す


                i=i+1
            else:
                comment.value = "エラー"
                break
        
                
    comment = ft.Text(color=ft.colors.RED)

    carsol_posi = ft.FilledTonalButton(text="カーソルの位置を取得", on_click=carsol_posi_clicked)

    second_W = ft.TextField(hint_text="待機時間", width=120)
    W_button = ft.FilledTonalButton(text="Add", on_click=W_button_click)

    A_number = ft.Dropdown(label="ツールバーのNo. ", width=120, options=[ft.dropdown.Option(1), ft.dropdown.Option(2), ft.dropdown.Option(3), ft.dropdown.Option(4), ft.dropdown.Option(5), ft.dropdown.Option(6), ft.dropdown.Option(7), ft.dropdown.Option(8), ft.dropdown.Option(9)])
    A_submi = ft.FilledTonalButton(text="Add", on_click=A_submi_click)

    mozi_field = ft.TextField(hint_text="入力したい文字")
    mozi_button = ft.FilledTonalButton(text="Add", on_click=mozi_button_click)

    def update_dd():
        move_posi.options = [ft.dropdown.Option(str(j)) for j in ryo]
        start_posi.options = [ft.dropdown.Option(str(j)) for j in ryo]
        end_posi.options = [ft.dropdown.Option(str(j)) for j in ryo]
    move_posi = ft.Dropdown(label="移動先No.",width=120, options=[])
    posi_button = ft.FilledTonalButton(text="Add", on_click=posi_button_click)

    lclick_button = ft.FilledTonalButton(text="左クリック", on_click=lclick_button_click)
    rclick_button = ft.FilledTonalButton(text="右クリック", on_click=rclick_button_click)
    ctrl_c = ft.FilledTonalButton(text="コピー", on_click=copy_click)
    ctrl_v = ft.FilledTonalButton(text="ペースト", on_click=paste_click)
    enter_button = ft.FilledTonalButton(text="エンター", on_click=enter_click)

    full_button = ft.FilledTonalButton(text="全画面", on_click=full_button_click)
    screenshot = ft.FilledTonalButton(text="スクリーンショット", on_click=screenshot_click)

    key_field = ft.TextField(hint_text="キー入力がありません", read_only=True)
    key_add = ft.FilledTonalButton(text="キー入力", on_click=lambda e: key_in(e, Keys))
    key_button = ft.FilledTonalButton(text="Add", on_click=lambda e: key_button_click(e, Keys))
    
    start_posi = ft.Dropdown(label="ドラッグ",width=120, options=[])
    end_posi = ft.Dropdown(label="ドロップ",width=120, options=[])
    drdr_button = ft.FilledTonalButton(text="Add", on_click=drdr_button_click)



    

    pick_files_dialog = ft.FilePicker(on_result=VBA_select_click)
    page.overlay.append(pick_files_dialog)
    pick_files_dialog2 = ft.FilePicker(on_result=excel_select_click)
    page.overlay.append(pick_files_dialog2)

    pick_files_PKL = ft.FilePicker(on_result=import_select_click)
    page.overlay.append(pick_files_PKL)
    pick_files_TXT = ft.FilePicker(on_result=txt_select_click)
    page.overlay.append(pick_files_TXT)


    VBA_select = ft.ElevatedButton(text="VBAを取得(txt)", icon = ft.icons.FILE_OPEN, on_click=lambda _: pick_files_dialog.pick_files(allow_multiple=False))
    excel_select = ft.ElevatedButton(text="excelのpathを取得(xlsm)", icon = ft.icons.FILE_OPEN, on_click=lambda _: pick_files_dialog2.pick_files(allow_multiple=False), visible=False, color=ft.colors.GREEN)

    chart_clear = ft.FilledTonalButton(text="clear",  on_click=clear_click)#clearのボタン未実装
    start_button = ft.FilledTonalButton(text="start", on_click=start_button_click)


    newxl_name = ft.TextField(hint_text="新しく作るExcelの名前")
    newxl_name_button = ft.FilledTonalButton(text="Add", on_click=newxl_ST)

    set_text = ft.TextField(hint_text="exportするファイルの名前", width=200)
    export_button = ft.FilledTonalButton(text="export", on_click=set_click)
    import_button = ft.ElevatedButton(text="import", icon = ft.icons.AIRPLAY, on_click=lambda _: pick_files_PKL.pick_files(allow_multiple=False))
    text_import = ft.ElevatedButton(text="txtからimport", icon = ft.icons.AD_UNITS_SHARP, on_click=lambda _: pick_files_TXT.pick_files(allow_multiple=False))


    #位置情報一覧
    locate_row = ft.Row()
    #チャート
    count_column = ft.ListView(expand=True)


    #UIの実装

    titlebar = ft.Container(ft.Row([
                ft.WindowDragArea(ft.Container(
                    ft.Text("AA", weight=ft.FontWeight.BOLD, color=ft.colors.ORANGE), 
                    padding=7,
                ),
                expand=True,
                ),
                ft.IconButton(ft.icons.CLOSE, icon_size=13, on_click=lambda _: page.window_close()),
            ],
            
            ),
            bgcolor=ft.colors.WHITE,   
        )
                

    header = ft.Column([ft.Row([ft.Text("コメント : ", color=ft.colors.BLACK), comment]), ft.Text("位置情報記録", color=ft.colors.BLUE), locate_row])

    left_side = ft.Container(ft.Column(
                    [
                        ft.Text("追加したい操作を選ぶ"),
                        ft.Row([carsol_posi]),
                        ft.Row([second_W, W_button, A_number, A_submi, move_posi, posi_button]),
                        ft.Row([lclick_button, rclick_button, ctrl_c, ctrl_v, enter_button]),
                        ft.Row([full_button, screenshot]),
                        ft.Row([key_field, key_add, key_button]),
                        ft.Row([mozi_field, mozi_button]),
                        ft.Row([start_posi, ft.Text("から"), end_posi, drdr_button]),

                        ft.Text("Excel関連", color=ft.colors.GREEN),
                        ft.Row([newxl_name, newxl_name_button]),
                        ft.Row([VBA_select, excel_select]),
                        ft.Row([set_text, export_button, import_button, text_import])
                    ],
                    scroll=ft.ScrollMode.AUTO,
                    expand=True
    ))

    count_column_list = ft.Container(
        content=count_column,
        border=ft.border.all(1, ft.colors.BLACK),  # 枠線の設定
        padding=10,  # 枠線とListViewの間に余白を追加
        expand=True  # ListViewが親コンテナに拡張されるようにする
    )

    start_clear = ft.Row([start_button, chart_clear])

    right_side = ft.Column([ft.Text("チャート"), count_column_list, start_clear],expand=True)

    row = ft.Row([left_side, right_side],expand=True)

    gamen = ft.Column([titlebar, header, row],expand=True)

    page.add(gamen)

    #初期設定ーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーー
    i=0
    while True:
            if i >= len(count):
                page.update()
                break
            elif count[i]=="TIME":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text(count[i]), ft.Text("秒待機")]))
                i=i+1
            elif count[i]=="TOOL":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text("ツールバーの"), ft.Text(count[i]), ft.Text("番目のアプリを選択")]))
                i=i+1
            elif count[i]=="MOZI":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text("文字列「"), ft.Text(count[i]), ft.Text("」を入力")]))
                i=i+1
            elif count[i]=="POSI":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text("カーソルを"), ft.Text(count[i]), ft.Text("に移動")]))
                i=i+1
            elif count[i]=="LCLICK":
                count_column.controls.append(ft.Row([ft.Text("左クリック")]))
                i=i+1
            elif count[i]=="RCLICK":
                count_column.controls.append(ft.Row([ft.Text("右クリック")]))
                i=i+1
            elif count[i]=="ENTER":
                count_column.controls.append(ft.Row([ft.Text("エンター")]))
                i=i+1
            elif count[i]=="DRDR":
                i=i+1
                stdr=count[i]
                i=i+1
                endr=count[i]
                count_column.controls.append(ft.Row([ft.Text(stdr), ft.Text("から"), ft.Text(endr), ft.Text("にドラッグドロップ")]))
                i=i+1
            elif count[i]=="COPY":
                count_column.controls.append(ft.Row([ft.Text("コピー")]))
                i=i+1
            elif count[i]=="PASTE":
                count_column.controls.append(ft.Row([ft.Text("ペースト")]))
                i=i+1
            elif count[i]=="FULL":
                count_column.controls.append(ft.Row([ft.Text("全画面表示")]))
                i=i+1
            elif count[i]=="SHOT":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text("スクリーンショット")]))
                i=i+1
            elif count[i]=="KEYCON":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text(count[i]), ft.Text("のキーを入力")]))
                i=i+1
            elif count[i]=="NEWXL":
                i=i+1
                count_column.controls.append(ft.Row([ft.Text(count[i]), ft.Text(".xlsxを作成")]))
                i=i+1
            elif count[i]=="VBA":
                i=i+1
                macro_path = count[i]
                i=i+1
                excel_path = count[i]
                count_column.controls.append(ft.Row([ft.Text(macro_path), ft.Text("を")]))
                count_column.controls.append(ft.Row([ft.Text(excel_path), ft.Text("に適応")]))
                i=i+1
            else:
                page.update()
                break





ft.app(target=main)